import csv
import io
from datetime import date

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .models import ActivityLog, BudgetLimit, Category, RecurringTransaction, SavingsGoal, Transaction, Wallet
from .views import generate_due_recurring_transactions

User = get_user_model()


class TransactionSecurityTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.other_user = User.objects.create_user(username="bob", email="bob@example.com", password="pass12345")
        self.own_category = Category.objects.create(user=self.user, name="Food")
        self.foreign_category = Category.objects.create(user=self.other_user, name="Rent")
        self.own_wallet = Wallet.objects.create(user=self.user, name="Cash")
        self.foreign_wallet = Wallet.objects.create(user=self.other_user, name="Card")
        self.client.force_login(self.user)

    def test_cannot_attach_transaction_to_another_users_category(self):
        response = self.client.post(reverse("add_expense"), {
            "amount": "50",
            "date": "2026-01-10",
            "category": self.foreign_category.id,
            "wallet": self.own_wallet.id,
            "note": "",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Transaction.objects.count(), 0)

    def test_cannot_attach_transaction_to_another_users_wallet(self):
        response = self.client.post(reverse("add_expense"), {
            "amount": "50",
            "date": "2026-01-10",
            "category": self.own_category.id,
            "wallet": self.foreign_wallet.id,
            "note": "",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Transaction.objects.count(), 0)

    def test_can_add_expense_with_own_category(self):
        response = self.client.post(reverse("add_expense"), {
            "amount": "50",
            "date": "2026-01-10",
            "category": self.own_category.id,
            "wallet": self.own_wallet.id,
            "note": "Groceries",
        })
        self.assertRedirects(response, reverse("dashboard"))
        tx = Transaction.objects.get()
        self.assertEqual(tx.user, self.user)
        self.assertEqual(tx.category, self.own_category)
        self.assertEqual(tx.wallet, self.own_wallet)
        self.assertEqual(tx.type, Transaction.EXPENSE)

    def test_rejects_non_numeric_amount(self):
        response = self.client.post(reverse("add_income"), {
            "amount": "not-a-number",
            "date": "2026-01-10",
            "category": self.own_category.id,
            "wallet": self.own_wallet.id,
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Transaction.objects.count(), 0)

    def test_rejects_negative_amount(self):
        response = self.client.post(reverse("add_income"), {
            "amount": "-10",
            "date": "2026-01-10",
            "category": self.own_category.id,
            "wallet": self.own_wallet.id,
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Transaction.objects.count(), 0)

    def test_rejects_invalid_date(self):
        response = self.client.post(reverse("add_income"), {
            "amount": "10",
            "date": "not-a-date",
            "category": self.own_category.id,
            "wallet": self.own_wallet.id,
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Transaction.objects.count(), 0)

    def test_form_values_are_preserved_after_a_validation_error(self):
        response = self.client.post(reverse("add_expense"), {
            "amount": "-10",
            "date": "2026-01-10",
            "category": self.own_category.id,
            "wallet": self.own_wallet.id,
            "note": "Coffee",
        })
        self.assertContains(response, 'value="-10"')
        self.assertContains(response, 'value="2026-01-10"')
        self.assertContains(response, 'value="Coffee"')
        self.assertContains(response, f'value="{self.own_category.id}" selected')
        self.assertContains(response, f'value="{self.own_wallet.id}" selected')


class CategoryFormTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.client.force_login(self.user)

    def test_duplicate_category_name_shows_error_instead_of_crashing(self):
        Category.objects.create(user=self.user, name="Food")
        response = self.client.post(reverse("category_add"), {"name": "Food", "icon": "", "color": ""})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Category.objects.filter(user=self.user).count(), 1)

    def test_blank_category_name_is_rejected(self):
        response = self.client.post(reverse("category_add"), {"name": "   ", "icon": "", "color": ""})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Category.objects.count(), 0)

    def test_color_field_renders_as_a_color_picker(self):
        response = self.client.get(reverse("category_add"))
        self.assertContains(response, 'type="color"')

    def test_color_field_has_a_sensible_default_for_new_categories(self):
        response = self.client.get(reverse("category_add"))
        self.assertEqual(response.context["form"].initial.get("color"), "#6366f1")

    def test_can_create_category_with_a_picked_color(self):
        response = self.client.post(reverse("category_add"), {"name": "Fun", "icon": "", "color": "#ff00aa"})
        self.assertRedirects(response, reverse("categories"))
        self.assertEqual(Category.objects.get(name="Fun").color, "#ff00aa")

    def test_accepts_a_multi_codepoint_emoji_as_icon(self):
        # A "single" emoji like a family or a flag can be several Unicode
        # code points long — this used to trip the old max_length=10 limit
        # with a confusing "has at most 10 characters" error.
        family_emoji = "👨‍👩‍👧‍👦"
        response = self.client.post(reverse("category_add"), {"name": "Family", "icon": family_emoji, "color": ""})
        self.assertRedirects(response, reverse("categories"))
        self.assertEqual(Category.objects.get(name="Family").icon, family_emoji)

    def test_overly_long_icon_gets_a_friendly_error_message(self):
        response = self.client.post(reverse("category_add"), {"name": "Long", "icon": "x" * 40, "color": ""})
        self.assertContains(response, "try a single emoji instead")
        self.assertEqual(Category.objects.filter(name="Long").count(), 0)


class TransactionsPaginationTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.category = Category.objects.create(user=self.user, name="Food")
        self.wallet = Wallet.objects.create(user=self.user, name="Cash")
        self.client.force_login(self.user)
        for i in range(60):
            Transaction.objects.create(
                user=self.user,
                type=Transaction.EXPENSE,
                amount=10,
                date=date(2026, 1, 1),
                category=self.category,
                wallet=self.wallet,
            )

    def test_list_is_paginated(self):
        response = self.client.get(reverse("transactions"))
        self.assertEqual(len(response.context["page_obj"]), 50)
        self.assertTrue(response.context["page_obj"].has_next())


class CompareViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.category = Category.objects.create(user=self.user, name="Food")
        self.wallet = Wallet.objects.create(user=self.user, name="Cash")
        self.client.force_login(self.user)
        Transaction.objects.create(user=self.user, type=Transaction.INCOME, amount=1000, date=date(2025, 1, 15), category=self.category, wallet=self.wallet)
        Transaction.objects.create(user=self.user, type=Transaction.EXPENSE, amount=300, date=date(2025, 1, 20), category=self.category, wallet=self.wallet)
        Transaction.objects.create(user=self.user, type=Transaction.INCOME, amount=1200, date=date(2026, 2, 5), category=self.category, wallet=self.wallet)
        Transaction.objects.create(user=self.user, type=Transaction.EXPENSE, amount=400, date=date(2026, 2, 10), category=self.category, wallet=self.wallet)

    def test_defaults_to_most_recent_year(self):
        response = self.client.get(reverse("compare"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_year"], 2026)
        self.assertEqual(response.context["year_total_income"], 1200)
        self.assertEqual(response.context["year_total_expense"], 400)

    def test_can_select_a_specific_year(self):
        response = self.client.get(reverse("compare"), {"year": "2025"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_year"], 2025)
        self.assertEqual(response.context["year_total_income"], 1000)
        self.assertEqual(response.context["year_total_expense"], 300)

    def test_year_over_year_totals(self):
        response = self.client.get(reverse("compare"))
        year_rows = {r["year"]: r for r in response.context["year_rows"]}
        self.assertEqual(year_rows[2025]["income"], 1000)
        self.assertEqual(year_rows[2026]["income"], 1200)

    def test_invalid_year_param_does_not_crash(self):
        response = self.client.get(reverse("compare"), {"year": "not-a-year"})
        self.assertEqual(response.status_code, 200)

    def test_only_sees_own_transactions(self):
        other = User.objects.create_user(username="bob", email="bob@example.com", password="pass12345")
        other_category = Category.objects.create(user=other, name="Rent")
        other_wallet = Wallet.objects.create(user=other, name="Card")
        Transaction.objects.create(user=other, type=Transaction.INCOME, amount=99999, date=date(2026, 3, 1), category=other_category, wallet=other_wallet)

        response = self.client.get(reverse("compare"), {"year": "2026"})
        self.assertEqual(response.context["year_total_income"], 1200)

    def test_shows_empty_state_when_no_data_for_selected_year(self):
        response = self.client.get(reverse("compare"), {"year": "2030"})
        self.assertContains(response, "No transactions in 2030 yet.")
        self.assertNotContains(response, 'id="compareChart"')

    def test_shows_chart_when_year_has_data(self):
        response = self.client.get(reverse("compare"), {"year": "2026"})
        self.assertContains(response, 'id="compareChart"')
        self.assertNotContains(response, "No transactions in 2026 yet.")


class ExportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.other_user = User.objects.create_user(username="bob", email="bob@example.com", password="pass12345")
        self.category = Category.objects.create(user=self.user, name="Food")
        self.wallet = Wallet.objects.create(user=self.user, name="Cash")
        other_category = Category.objects.create(user=self.other_user, name="Secret")
        other_wallet = Wallet.objects.create(user=self.other_user, name="Card")
        self.client.force_login(self.user)
        Transaction.objects.create(user=self.user, type=Transaction.INCOME, amount=500, date=date(2026, 1, 10), category=self.category, wallet=self.wallet, note="Salary")
        Transaction.objects.create(user=self.user, type=Transaction.EXPENSE, amount=75, date=date(2026, 1, 12), category=self.category, wallet=self.wallet, note="Groceries")
        Transaction.objects.create(user=self.other_user, type=Transaction.INCOME, amount=9999, date=date(2026, 1, 10), category=other_category, wallet=other_wallet, note="Not mine")

    def test_csv_export_contains_only_own_rows(self):
        response = self.client.get(reverse("export_transactions_csv"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        rows = list(csv.reader(io.StringIO(response.content.decode())))
        self.assertEqual(rows[0], ["Date", "Type", "Category", "Wallet", "Amount", "Currency", "Note"])
        notes = [r[6] for r in rows[1:]]
        self.assertIn("Salary", notes)
        self.assertIn("Groceries", notes)
        self.assertNotIn("Not mine", notes)

    def test_csv_export_respects_type_filter(self):
        response = self.client.get(reverse("export_transactions_csv"), {"type": "income"})
        rows = list(csv.reader(io.StringIO(response.content.decode())))
        notes = [r[6] for r in rows[1:]]
        self.assertEqual(notes, ["Salary"])

    def test_csv_export_respects_wallet_filter(self):
        other_wallet = Wallet.objects.create(user=self.user, name="Savings")
        Transaction.objects.create(user=self.user, type=Transaction.EXPENSE, amount=20, date=date(2026, 1, 15), category=self.category, wallet=other_wallet, note="From savings")

        response = self.client.get(reverse("export_transactions_csv"), {"wallet": self.wallet.id})
        rows = list(csv.reader(io.StringIO(response.content.decode())))
        notes = [r[6] for r in rows[1:]]
        self.assertIn("Salary", notes)
        self.assertNotIn("From savings", notes)

    def test_xlsx_export_contains_only_own_rows(self):
        response = self.client.get(reverse("export_transactions_xlsx"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        notes = [row[6].value for row in ws.iter_rows(min_row=2)]
        self.assertIn("Salary", notes)
        self.assertIn("Groceries", notes)
        self.assertNotIn("Not mine", notes)

    def test_pdf_export_returns_pdf(self):
        response = self.client.get(reverse("export_transactions_pdf"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_export_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse("export_transactions_csv"))
        self.assertEqual(response.status_code, 302)


class EmptyCategoriesStateTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.client.force_login(self.user)

    def test_add_income_shows_create_category_prompt_when_no_categories(self):
        response = self.client.get(reverse("add_income"))
        self.assertContains(response, "You don't have any categories yet")
        self.assertContains(response, reverse("category_add"))
        self.assertNotContains(response, "<select name=\"category\"")

    def test_add_expense_shows_create_category_prompt_when_no_categories(self):
        response = self.client.get(reverse("add_expense"))
        self.assertContains(response, "You don't have any categories yet")
        self.assertNotContains(response, "<select name=\"category\"")

    def test_add_income_shows_wallet_prompt_once_a_category_exists_but_no_wallet(self):
        Category.objects.create(user=self.user, name="Salary")
        response = self.client.get(reverse("add_income"))
        self.assertNotContains(response, "You don't have any categories yet")
        self.assertContains(response, "You don't have any wallets yet")
        self.assertContains(response, reverse("wallet_add"))
        self.assertNotContains(response, "<select name=\"category\"")

    def test_add_income_shows_form_once_a_category_and_wallet_exist(self):
        Category.objects.create(user=self.user, name="Salary")
        Wallet.objects.create(user=self.user, name="Cash")
        response = self.client.get(reverse("add_income"))
        self.assertNotContains(response, "You don't have any categories yet")
        self.assertNotContains(response, "You don't have any wallets yet")
        self.assertContains(response, "<select name=\"category\"")
        self.assertContains(response, "<select name=\"wallet\"")


class BudgetLimitFormTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.category = Category.objects.create(user=self.user, name="Food")
        self.client.force_login(self.user)

    def test_month_field_renders_as_a_month_picker(self):
        response = self.client.get(reverse("budget_add"))
        self.assertContains(response, 'type="month"')

    def test_can_create_a_budget_limit_with_month_only_value(self):
        response = self.client.post(reverse("budget_add"), {
            "category": self.category.id,
            "month": "2026-03",
            "limit": "500",
        })
        self.assertRedirects(response, reverse("budgets"))
        b = BudgetLimit.objects.get()
        self.assertEqual(b.month, date(2026, 3, 1))
        self.assertEqual(b.user, self.user)

    def test_edit_page_shows_existing_month_in_picker_format(self):
        b = BudgetLimit.objects.create(user=self.user, category=self.category, month=date(2026, 5, 1), limit=300)
        response = self.client.get(reverse("budget_edit", args=[b.id]))
        self.assertContains(response, 'value="2026-05"')


class BudgetsViewNoLimitStateTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        Category.objects.create(user=self.user, name="Food")
        self.client.force_login(self.user)

    def test_shows_explicit_no_limit_message_instead_of_hiding_the_row(self):
        response = self.client.get(reverse("budgets"))
        self.assertContains(response, "No limit set for this category yet.")


class SavingsGoalTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.other_user = User.objects.create_user(username="bob", email="bob@example.com", password="pass12345")
        self.client.force_login(self.user)

    def test_can_create_a_goal(self):
        response = self.client.post(reverse("savings_goal_add"), {
            "name": "Emergency fund",
            "target_amount": "1000",
            "target_date": "",
        })
        self.assertRedirects(response, reverse("savings_goals"))
        goal = SavingsGoal.objects.get()
        self.assertEqual(goal.user, self.user)
        self.assertEqual(goal.saved_amount, 0)

    def test_rejects_zero_or_negative_target(self):
        response = self.client.post(reverse("savings_goal_add"), {
            "name": "Bad goal", "target_amount": "0", "target_date": "",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(SavingsGoal.objects.count(), 0)

    def test_add_funds_increases_saved_amount(self):
        goal = SavingsGoal.objects.create(user=self.user, name="Trip", target_amount=500)
        self.client.post(reverse("savings_goal_add_funds", args=[goal.id]), {"amount": "120"})
        goal.refresh_from_db()
        self.assertEqual(goal.saved_amount, 120)

        self.client.post(reverse("savings_goal_add_funds", args=[goal.id]), {"amount": "30"})
        goal.refresh_from_db()
        self.assertEqual(goal.saved_amount, 150)

    def test_add_funds_rejects_invalid_amount(self):
        goal = SavingsGoal.objects.create(user=self.user, name="Trip", target_amount=500)
        self.client.post(reverse("savings_goal_add_funds", args=[goal.id]), {"amount": "not-a-number"})
        goal.refresh_from_db()
        self.assertEqual(goal.saved_amount, 0)

    def test_progress_percentage_is_capped_at_100(self):
        goal = SavingsGoal.objects.create(user=self.user, name="Overfunded", target_amount=100, saved_amount=250)
        response = self.client.get(reverse("savings_goals"))
        rows = response.context["rows"]
        self.assertEqual(rows[0]["pct"], 100)
        self.assertTrue(rows[0]["is_complete"])

    def test_cannot_add_funds_to_another_users_goal(self):
        other_goal = SavingsGoal.objects.create(user=self.other_user, name="Not mine", target_amount=500)
        response = self.client.post(reverse("savings_goal_add_funds", args=[other_goal.id]), {"amount": "100"})
        self.assertEqual(response.status_code, 404)

    def test_cannot_delete_another_users_goal(self):
        other_goal = SavingsGoal.objects.create(user=self.other_user, name="Not mine", target_amount=500)
        response = self.client.post(reverse("savings_goal_delete", args=[other_goal.id]))
        self.assertEqual(response.status_code, 404)
        self.assertTrue(SavingsGoal.objects.filter(pk=other_goal.pk).exists())

    def test_delete_removes_the_goal(self):
        goal = SavingsGoal.objects.create(user=self.user, name="Trip", target_amount=500)
        response = self.client.post(reverse("savings_goal_delete", args=[goal.id]))
        self.assertRedirects(response, reverse("savings_goals"))
        self.assertFalse(SavingsGoal.objects.filter(pk=goal.pk).exists())

    def test_goals_list_shows_empty_state(self):
        response = self.client.get(reverse("savings_goals"))
        self.assertContains(response, "No savings goals yet")


class ActivityLogTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.other_user = User.objects.create_user(username="bob", email="bob@example.com", password="pass12345")
        self.category = Category.objects.create(user=self.user, name="Food")
        self.wallet = Wallet.objects.create(user=self.user, name="Cash")
        self.client.force_login(self.user)

    def test_adding_a_transaction_is_logged(self):
        self.client.post(reverse("add_expense"), {
            "amount": "50", "date": "2026-01-10", "category": self.category.id, "wallet": self.wallet.id, "note": "Groceries",
        })
        entry = ActivityLog.objects.get()
        self.assertEqual(entry.user, self.user)
        self.assertEqual(entry.action, ActivityLog.CREATE)
        self.assertEqual(entry.model_name, "Transaction")
        self.assertIn("Food", entry.object_repr)

    def test_editing_a_transaction_is_logged(self):
        tx = Transaction.objects.create(user=self.user, type=Transaction.EXPENSE, amount=50, date="2026-01-10", category=self.category, wallet=self.wallet)
        self.client.post(reverse("transaction_edit", args=[tx.id]), {
            "type": "expense", "amount": "75", "date": "2026-01-10", "category": self.category.id, "wallet": self.wallet.id, "note": "",
        })
        entry = ActivityLog.objects.filter(action=ActivityLog.UPDATE).get()
        self.assertEqual(entry.model_name, "Transaction")

    def test_deleting_a_transaction_is_logged(self):
        tx = Transaction.objects.create(user=self.user, type=Transaction.EXPENSE, amount=50, date="2026-01-10", category=self.category, wallet=self.wallet)
        self.client.post(reverse("transaction_delete", args=[tx.id]))
        entry = ActivityLog.objects.filter(action=ActivityLog.DELETE).get()
        self.assertEqual(entry.model_name, "Transaction")

    def test_budget_limit_changes_are_logged(self):
        self.client.post(reverse("budget_add"), {
            "category": self.category.id, "month": "2026-01", "limit": "300",
        })
        entry = ActivityLog.objects.get()
        self.assertEqual(entry.model_name, "Budget Limit")
        self.assertEqual(entry.action, ActivityLog.CREATE)

    def test_activity_log_only_shows_own_entries(self):
        ActivityLog.objects.create(user=self.other_user, action=ActivityLog.CREATE, model_name="Transaction", object_repr="Not mine")
        ActivityLog.objects.create(user=self.user, action=ActivityLog.CREATE, model_name="Transaction", object_repr="Mine")
        response = self.client.get(reverse("activity_log"))
        self.assertContains(response, "Mine")
        self.assertNotContains(response, "Not mine")

    def test_activity_log_page_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse("activity_log"))
        self.assertEqual(response.status_code, 302)


class WalletTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.other_user = User.objects.create_user(username="bob", email="bob@example.com", password="pass12345")
        self.category = Category.objects.create(user=self.user, name="Food")
        self.client.force_login(self.user)

    def test_can_create_a_wallet(self):
        response = self.client.post(reverse("wallet_add"), {"name": "Card", "icon": "💳"})
        self.assertRedirects(response, reverse("wallets"))
        wallet = Wallet.objects.get()
        self.assertEqual(wallet.user, self.user)
        self.assertEqual(wallet.name, "Card")

    def test_duplicate_wallet_name_is_rejected(self):
        Wallet.objects.create(user=self.user, name="Cash")
        response = self.client.post(reverse("wallet_add"), {"name": "Cash", "icon": ""})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Wallet.objects.filter(user=self.user).count(), 1)

    def test_blank_wallet_name_is_rejected(self):
        response = self.client.post(reverse("wallet_add"), {"name": "   ", "icon": ""})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Wallet.objects.count(), 0)

    def test_wallets_list_shows_empty_state(self):
        response = self.client.get(reverse("wallets"))
        self.assertContains(response, "No wallets yet")

    def test_wallet_balance_reflects_its_own_transactions_only(self):
        cash = Wallet.objects.create(user=self.user, name="Cash")
        card = Wallet.objects.create(user=self.user, name="Card")
        Transaction.objects.create(user=self.user, type=Transaction.INCOME, amount=1000, date="2026-01-01", category=self.category, wallet=cash)
        Transaction.objects.create(user=self.user, type=Transaction.EXPENSE, amount=200, date="2026-01-02", category=self.category, wallet=cash)
        Transaction.objects.create(user=self.user, type=Transaction.INCOME, amount=500, date="2026-01-03", category=self.category, wallet=card)

        response = self.client.get(reverse("wallets"))
        balances = {r["wallet"].id: r["balance"] for r in response.context["rows"]}
        self.assertEqual(balances[cash.id], 800)
        self.assertEqual(balances[card.id], 500)

    def test_only_own_wallets_appear_in_add_transaction_dropdown(self):
        own_wallet = Wallet.objects.create(user=self.user, name="Cash")
        Wallet.objects.create(user=self.other_user, name="Not mine")
        response = self.client.get(reverse("add_income"))
        self.assertContains(response, "Cash")
        self.assertNotContains(response, "Not mine")


class RecurringTransactionGenerationTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.category = Category.objects.create(user=self.user, name="Subscriptions")
        self.wallet = Wallet.objects.create(user=self.user, name="Card")
        self.client.force_login(self.user)

    def test_generates_first_occurrence_in_start_month_when_day_already_passed_today(self):
        RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.EXPENSE, amount=15,
            category=self.category, wallet=self.wallet,
            day_of_month=5, start_date=date(2026, 1, 1),
        )
        generate_due_recurring_transactions(self.user, today=date(2026, 1, 10))
        tx = Transaction.objects.get()
        self.assertEqual(tx.date, date(2026, 1, 5))
        self.assertEqual(tx.amount, 15)
        self.assertEqual(tx.wallet, self.wallet)

    def test_does_not_generate_before_the_day_of_month_arrives(self):
        RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.EXPENSE, amount=15,
            category=self.category, wallet=self.wallet,
            day_of_month=20, start_date=date(2026, 1, 1),
        )
        generate_due_recurring_transactions(self.user, today=date(2026, 1, 10))
        self.assertEqual(Transaction.objects.count(), 0)

    def test_does_not_generate_before_start_date(self):
        RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.EXPENSE, amount=15,
            category=self.category, wallet=self.wallet,
            day_of_month=1, start_date=date(2026, 6, 1),
        )
        generate_due_recurring_transactions(self.user, today=date(2026, 3, 1))
        self.assertEqual(Transaction.objects.count(), 0)

    def test_catches_up_multiple_missed_months(self):
        RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.INCOME, amount=2000,
            category=self.category, wallet=self.wallet,
            day_of_month=1, start_date=date(2026, 1, 1),
        )
        generate_due_recurring_transactions(self.user, today=date(2026, 4, 15))
        dates = sorted(Transaction.objects.values_list("date", flat=True))
        self.assertEqual(dates, [date(2026, 1, 1), date(2026, 2, 1), date(2026, 3, 1), date(2026, 4, 1)])

    def test_does_not_duplicate_on_repeated_calls(self):
        r = RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.EXPENSE, amount=15,
            category=self.category, wallet=self.wallet,
            day_of_month=5, start_date=date(2026, 1, 1),
        )
        generate_due_recurring_transactions(self.user, today=date(2026, 1, 10))
        generate_due_recurring_transactions(self.user, today=date(2026, 1, 10))
        generate_due_recurring_transactions(self.user, today=date(2026, 1, 31))
        self.assertEqual(Transaction.objects.count(), 1)
        r.refresh_from_db()
        self.assertEqual(r.last_generated, date(2026, 1, 5))

    def test_paused_recurring_transaction_does_not_generate(self):
        RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.EXPENSE, amount=15,
            category=self.category, wallet=self.wallet,
            day_of_month=5, start_date=date(2026, 1, 1), is_active=False,
        )
        generate_due_recurring_transactions(self.user, today=date(2026, 2, 1))
        self.assertEqual(Transaction.objects.count(), 0)

    def test_day_28_works_correctly_in_february(self):
        RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.EXPENSE, amount=10,
            category=self.category, wallet=self.wallet,
            day_of_month=28, start_date=date(2026, 1, 1),
        )
        generate_due_recurring_transactions(self.user, today=date(2026, 3, 1))
        dates = sorted(Transaction.objects.values_list("date", flat=True))
        self.assertEqual(dates, [date(2026, 1, 28), date(2026, 2, 28)])

    def test_generation_is_scoped_to_the_given_user(self):
        other = User.objects.create_user(username="bob", email="bob@example.com", password="pass12345")
        other_category = Category.objects.create(user=other, name="Rent")
        other_wallet = Wallet.objects.create(user=other, name="Card")
        RecurringTransaction.objects.create(
            user=other, type=RecurringTransaction.EXPENSE, amount=999,
            category=other_category, wallet=other_wallet,
            day_of_month=1, start_date=date(2026, 1, 1),
        )
        generate_due_recurring_transactions(self.user, today=date(2026, 2, 1))
        self.assertEqual(Transaction.objects.count(), 0)

    def test_generated_transaction_is_logged(self):
        RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.EXPENSE, amount=15,
            category=self.category, wallet=self.wallet,
            day_of_month=5, start_date=date(2026, 1, 1),
        )
        generate_due_recurring_transactions(self.user, today=date(2026, 1, 10))
        entry = ActivityLog.objects.get()
        self.assertEqual(entry.model_name, "Transaction")
        self.assertIn("recurring", entry.object_repr)

    def test_dashboard_visit_triggers_generation(self):
        RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.INCOME, amount=3000,
            category=self.category, wallet=self.wallet,
            day_of_month=1, start_date=date(2020, 1, 1),
        )
        self.client.get(reverse("dashboard"))
        self.assertTrue(Transaction.objects.exists())


class RecurringTransactionCrudTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="alice", email="alice@example.com", password="pass12345")
        self.other_user = User.objects.create_user(username="bob", email="bob@example.com", password="pass12345")
        self.category = Category.objects.create(user=self.user, name="Subscriptions")
        self.wallet = Wallet.objects.create(user=self.user, name="Card")
        self.client.force_login(self.user)

    def test_add_prompts_for_category_when_none_exist(self):
        Category.objects.filter(user=self.user).delete()
        response = self.client.get(reverse("recurring_add"))
        self.assertContains(response, "You don't have any categories yet")

    def test_add_prompts_for_wallet_when_none_exist(self):
        Wallet.objects.filter(user=self.user).delete()
        response = self.client.get(reverse("recurring_add"))
        self.assertContains(response, "You don't have any wallets yet")

    def test_can_create_a_recurring_transaction(self):
        response = self.client.post(reverse("recurring_add"), {
            "type": "expense", "amount": "9.99", "category": self.category.id, "wallet": self.wallet.id,
            "day_of_month": "15", "start_date": "2026-01-01", "note": "Netflix", "is_active": "on",
        })
        self.assertRedirects(response, reverse("recurring_transactions"))
        r = RecurringTransaction.objects.get()
        self.assertEqual(r.user, self.user)
        self.assertEqual(r.day_of_month, 15)

    def test_rejects_zero_amount(self):
        response = self.client.post(reverse("recurring_add"), {
            "type": "expense", "amount": "0", "category": self.category.id, "wallet": self.wallet.id,
            "day_of_month": "15", "start_date": "2026-01-01",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(RecurringTransaction.objects.count(), 0)

    def test_rejects_day_of_month_above_28(self):
        response = self.client.post(reverse("recurring_add"), {
            "type": "expense", "amount": "10", "category": self.category.id, "wallet": self.wallet.id,
            "day_of_month": "31", "start_date": "2026-01-01",
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(RecurringTransaction.objects.count(), 0)

    def test_cannot_edit_another_users_recurring_transaction(self):
        other_category = Category.objects.create(user=self.other_user, name="Rent")
        other_wallet = Wallet.objects.create(user=self.other_user, name="Card")
        other_r = RecurringTransaction.objects.create(
            user=self.other_user, type=RecurringTransaction.EXPENSE, amount=10,
            category=other_category, wallet=other_wallet, day_of_month=1, start_date=date(2026, 1, 1),
        )
        response = self.client.get(reverse("recurring_edit", args=[other_r.id]))
        self.assertEqual(response.status_code, 404)

    def test_cannot_delete_another_users_recurring_transaction(self):
        other_category = Category.objects.create(user=self.other_user, name="Rent")
        other_wallet = Wallet.objects.create(user=self.other_user, name="Card")
        other_r = RecurringTransaction.objects.create(
            user=self.other_user, type=RecurringTransaction.EXPENSE, amount=10,
            category=other_category, wallet=other_wallet, day_of_month=1, start_date=date(2026, 1, 1),
        )
        response = self.client.post(reverse("recurring_delete", args=[other_r.id]))
        self.assertEqual(response.status_code, 404)
        self.assertTrue(RecurringTransaction.objects.filter(pk=other_r.pk).exists())

    def test_delete_removes_it(self):
        r = RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.EXPENSE, amount=10,
            category=self.category, wallet=self.wallet, day_of_month=1, start_date=date(2026, 1, 1),
        )
        response = self.client.post(reverse("recurring_delete", args=[r.id]))
        self.assertRedirects(response, reverse("recurring_transactions"))
        self.assertFalse(RecurringTransaction.objects.filter(pk=r.pk).exists())

    def test_list_shows_empty_state(self):
        response = self.client.get(reverse("recurring_transactions"))
        self.assertContains(response, "No recurring transactions yet")

    def test_management_command_generates_for_all_users(self):
        from io import StringIO
        from django.core.management import call_command

        RecurringTransaction.objects.create(
            user=self.user, type=RecurringTransaction.INCOME, amount=3000,
            category=self.category, wallet=self.wallet, day_of_month=1, start_date=date(2020, 1, 1),
        )
        out = StringIO()
        call_command("generate_recurring_transactions", stdout=out)
        self.assertTrue(Transaction.objects.exists())
        self.assertIn("Checked recurring transactions", out.getvalue())
