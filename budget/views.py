import csv
from calendar import monthrange
from datetime import date
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Sum
from django.db.models.functions import TruncMonth, TruncYear
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme

from .forms import BudgetLimitForm, CategoryForm, RecurringTransactionForm, SavingsGoalForm, TransactionForm, WalletForm
from .models import ActivityLog, BudgetLimit, Category, RecurringTransaction, SavingsGoal, Transaction, Wallet


def _safe_next_url(request, next_url):
    if next_url and url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        return next_url
    return None


def _log_activity(user, action, model_name, object_repr):
    ActivityLog.objects.create(
        user=user,
        action=action,
        model_name=model_name,
        object_repr=object_repr[:255],
    )


def _transaction_repr(tx, currency):
    return f"{tx.get_type_display()} · {tx.category.name} · {tx.wallet.name} · {tx.amount} {currency}"


def _budget_limit_repr(b, currency):
    return f"{b.category.name} · {b.month:%b %Y} limit · {b.limit} {currency}"


def _next_occurrence(d, day_of_month):
    """The same day_of_month, one month after d. day_of_month is capped at
    28 (see RecurringTransaction.day_of_month), so this never lands on an
    invalid date."""
    month = d.month + 1
    year = d.year + (month - 1) // 12
    month = (month - 1) % 12 + 1
    return date(year, month, day_of_month)


def generate_due_recurring_transactions(user, today=None):
    """Create real Transaction rows for any recurring transaction whose next
    occurrence has arrived. Catches up on every month missed since the last
    generated occurrence (e.g. if the user hasn't opened the app in a
    while), not just the most recent one."""
    today = today or date.today()
    currency = user.currency

    for r in RecurringTransaction.objects.filter(user=user, is_active=True, start_date__lte=today):
        if r.last_generated:
            candidate = _next_occurrence(r.last_generated, r.day_of_month)
        else:
            candidate = date(r.start_date.year, r.start_date.month, r.day_of_month)
            if candidate < r.start_date:
                candidate = _next_occurrence(candidate, r.day_of_month)

        while candidate <= today:
            tx = Transaction.objects.create(
                user=user,
                type=r.type,
                amount=r.amount,
                date=candidate,
                note=r.note,
                category=r.category,
                wallet=r.wallet,
            )
            _log_activity(
                user, ActivityLog.CREATE, "Transaction",
                f"{_transaction_repr(tx, currency)} (recurring)",
            )
            r.last_generated = candidate
            r.save(update_fields=["last_generated"])
            candidate = _next_occurrence(candidate, r.day_of_month)


@login_required
def dashboard_view(request):
    generate_due_recurring_transactions(request.user)

    transactions = Transaction.objects.filter(user=request.user)

    totals = transactions.aggregate(
        income=Sum("amount", filter=models.Q(type=Transaction.INCOME)),
        expense=Sum("amount", filter=models.Q(type=Transaction.EXPENSE)),
    )
    total_income = totals["income"] or Decimal("0")
    total_expense = totals["expense"] or Decimal("0")
    balance = total_income - total_expense

    last_transactions = transactions.select_related("category").order_by("-date")[:5]

    today = date.today()
    budget_alerts = [
        r for r in _budget_status_rows(request.user, _month_start(today), _month_end(today))
        if r["is_over"] or r["is_near"]
    ]

    context = {
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": balance,
        "transactions": last_transactions,
        "currency": request.user.currency,
        "budget_alerts": budget_alerts,
    }

    return render(request, "dashboard.html", context)


def _parse_amount(raw_amount):
    try:
        amount = Decimal(str(raw_amount).strip())
    except (InvalidOperation, ValueError, AttributeError):
        return None
    if amount <= 0:
        return None
    return amount


def _parse_date(raw_date):
    try:
        return date.fromisoformat(str(raw_date).strip())
    except (TypeError, ValueError):
        return None


def _add_transaction_view(request, tx_type, template_name):
    categories = Category.objects.filter(user=request.user).order_by("name")
    wallets = Wallet.objects.filter(user=request.user).order_by("name")

    if request.method == "POST":
        amount = _parse_amount(request.POST.get("amount"))
        tx_date = _parse_date(request.POST.get("date"))
        category = categories.filter(pk=request.POST.get("category")).first()
        wallet = wallets.filter(pk=request.POST.get("wallet")).first()
        note = request.POST.get("note", "").strip()

        errors = []
        if amount is None:
            errors.append("Please enter a valid amount greater than zero.")
        if tx_date is None:
            errors.append("Please enter a valid date.")
        if category is None:
            errors.append("Please choose one of your categories.")
        if wallet is None:
            errors.append("Please choose one of your wallets.")

        invalid_form_response = lambda: render(request, template_name, {
            "categories": categories,
            "wallets": wallets,
            "values": {
                "amount": request.POST.get("amount", ""),
                "date": request.POST.get("date", ""),
                "category": request.POST.get("category", ""),
                "wallet": request.POST.get("wallet", ""),
                "note": note,
            },
        })

        if errors:
            for error in errors:
                messages.error(request, error)
            return invalid_form_response()

        tx = Transaction(
            user=request.user,
            type=tx_type,
            amount=amount,
            date=tx_date,
            note=note,
            category=category,
            wallet=wallet,
        )
        try:
            tx.full_clean()
        except ValidationError as exc:
            for error in exc.messages:
                messages.error(request, error)
            return invalid_form_response()

        tx.save()
        _log_activity(request.user, ActivityLog.CREATE, "Transaction", _transaction_repr(tx, request.user.currency))
        messages.success(request, "Transaction saved.")
        return redirect("dashboard")

    return render(request, template_name, {"categories": categories, "wallets": wallets})


@login_required
def add_income_view(request):
    return _add_transaction_view(request, Transaction.INCOME, "income_add.html")


@login_required
def add_expense_view(request):
    return _add_transaction_view(request, Transaction.EXPENSE, "expense_add.html")


@login_required
def categories_view(request):
    cats = Category.objects.filter(user=request.user).order_by("name")
    return render(request, "categories.html", {"categories": cats})


@login_required
def category_add_view(request):
    next_url = _safe_next_url(request, request.POST.get("next") or request.GET.get("next"))

    if request.method == "POST":
        form = CategoryForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Category created!")
            return redirect(next_url) if next_url else redirect("categories")
    else:
        form = CategoryForm(user=request.user)

    return render(request, "category_add.html", {"form": form, "next": next_url, "mode": "add"})


@login_required
def category_edit_view(request, pk):
    category = get_object_or_404(Category, pk=pk, user=request.user)

    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Category updated.")
            return redirect("categories")
    else:
        form = CategoryForm(instance=category, user=request.user)

    return render(request, "category_add.html", {"form": form, "mode": "edit", "category": category})


@login_required
def category_delete_view(request, pk):
    category = get_object_or_404(Category, pk=pk, user=request.user)

    if request.method == "POST":
        category.delete()
        messages.success(request, "Category deleted.")
        return redirect("categories")

    return render(request, "category_confirm_delete.html", {
        "category": category,
        "tx_count": category.transactions.count(),
    })


@login_required
def wallets_view(request):
    wallets = Wallet.objects.filter(user=request.user).order_by("name")

    balances = {
        row["wallet_id"]: (row["income"] or Decimal("0")) - (row["expense"] or Decimal("0"))
        for row in (
            Transaction.objects.filter(user=request.user)
            .values("wallet_id")
            .annotate(
                income=Sum("amount", filter=models.Q(type=Transaction.INCOME)),
                expense=Sum("amount", filter=models.Q(type=Transaction.EXPENSE)),
            )
        )
    }

    rows = [{"wallet": w, "balance": balances.get(w.id, Decimal("0"))} for w in wallets]

    return render(request, "wallets.html", {"rows": rows, "currency": request.user.currency})


@login_required
def wallet_add_view(request):
    next_url = _safe_next_url(request, request.POST.get("next") or request.GET.get("next"))

    if request.method == "POST":
        form = WalletForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Wallet created!")
            return redirect(next_url) if next_url else redirect("wallets")
    else:
        form = WalletForm(user=request.user)

    return render(request, "wallet_add.html", {"form": form, "next": next_url, "mode": "add"})


@login_required
def wallet_edit_view(request, pk):
    wallet = get_object_or_404(Wallet, pk=pk, user=request.user)

    if request.method == "POST":
        form = WalletForm(request.POST, instance=wallet, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Wallet updated.")
            return redirect("wallets")
    else:
        form = WalletForm(instance=wallet, user=request.user)

    return render(request, "wallet_add.html", {"form": form, "mode": "edit", "wallet": wallet})


@login_required
def wallet_delete_view(request, pk):
    wallet = get_object_or_404(Wallet, pk=pk, user=request.user)

    if request.method == "POST":
        wallet.delete()
        messages.success(request, "Wallet deleted.")
        return redirect("wallets")

    return render(request, "wallet_confirm_delete.html", {
        "wallet": wallet,
        "tx_count": wallet.transactions.count(),
    })


def _filtered_transactions(request):
    """Shared filter parsing for the transactions list and the export views,
    so both stay in sync with the same type/category/wallet/date-range filters."""
    qs = Transaction.objects.filter(user=request.user).select_related("category", "wallet")
    t_type = request.GET.get("type", "").strip()
    cat_id = request.GET.get("category", "").strip()
    wallet_id = request.GET.get("wallet", "").strip()
    date_from = request.GET.get("from", "").strip()
    date_to = request.GET.get("to", "").strip()

    if t_type in (Transaction.INCOME, Transaction.EXPENSE):
        qs = qs.filter(type=t_type)

    if cat_id.isdigit():
        qs = qs.filter(category_id=int(cat_id))

    if wallet_id.isdigit():
        qs = qs.filter(wallet_id=int(wallet_id))

    if date_from:
        qs = qs.filter(date__gte=date_from)

    if date_to:
        qs = qs.filter(date__lte=date_to)

    filters = {"type": t_type, "category": cat_id, "wallet": wallet_id, "from": date_from, "to": date_to}
    return qs, filters


@login_required
def transactions_list_view(request):
    qs, filters = _filtered_transactions(request)

    total_income = qs.filter(type=Transaction.INCOME).aggregate(Sum("amount"))["amount__sum"] or 0
    total_expense = qs.filter(type=Transaction.EXPENSE).aggregate(Sum("amount"))["amount__sum"] or 0
    balance = total_income - total_expense

    categories = Category.objects.filter(user=request.user).order_by("name")
    wallets = Wallet.objects.filter(user=request.user).order_by("name")

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "transactions": page_obj,
        "page_obj": page_obj,
        "categories": categories,
        "wallets": wallets,
        "filters": filters,
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": balance,
        "currency": request.user.currency,
    }
    return render(request, "transactions_list.html", context)


_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _spreadsheet_safe(value):
    """Neutralize CSV/Excel formula injection: a category/wallet name or
    note starting with =, +, -, @ (or a tab/CR) would otherwise be
    interpreted as a formula by Excel/LibreOffice when the export is
    opened, letting user-controlled text run arbitrary spreadsheet
    formulas (including DDE) on whoever downloads the file."""
    text = "" if value is None else str(value)
    if text.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + text
    return text


@login_required
def export_transactions_csv_view(request):
    qs, _ = _filtered_transactions(request)
    qs = qs.order_by("-date")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="transactions.csv"'

    writer = csv.writer(response)
    writer.writerow(["Date", "Type", "Category", "Wallet", "Amount", "Currency", "Note"])
    for t in qs:
        writer.writerow([
            t.date.isoformat(),
            t.get_type_display(),
            _spreadsheet_safe(t.category.name),
            _spreadsheet_safe(t.wallet.name),
            t.amount,
            request.user.currency,
            _spreadsheet_safe(t.note),
        ])
    return response


@login_required
def export_transactions_xlsx_view(request):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    qs, _ = _filtered_transactions(request)
    qs = qs.order_by("-date")

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["Date", "Type", "Category", "Wallet", "Amount", "Currency", "Note"])
    for t in qs:
        ws.append([
            t.date.isoformat(),
            t.get_type_display(),
            _spreadsheet_safe(t.category.name),
            _spreadsheet_safe(t.wallet.name),
            float(t.amount),
            request.user.currency,
            _spreadsheet_safe(t.note),
        ])
    for i, width in enumerate([12, 10, 24, 18, 12, 10, 36], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="transactions.xlsx"'
    wb.save(response)
    return response


@login_required
def export_transactions_pdf_view(request):
    from weasyprint import HTML

    qs, _ = _filtered_transactions(request)
    qs = qs.order_by("-date")

    total_income = qs.filter(type=Transaction.INCOME).aggregate(Sum("amount"))["amount__sum"] or 0
    total_expense = qs.filter(type=Transaction.EXPENSE).aggregate(Sum("amount"))["amount__sum"] or 0

    html_string = render_to_string("exports/transactions_pdf.html", {
        "transactions": qs,
        "currency": request.user.currency,
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": total_income - total_expense,
        "generated_at": timezone.now(),
        "username": request.user.username,
    })

    pdf_bytes = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="transactions.pdf"'
    return response


@login_required
def transaction_edit_view(request, pk):
    tx = get_object_or_404(Transaction, pk=pk, user=request.user)

    if request.method == "POST":
        form = TransactionForm(request.POST, instance=tx, user=request.user)
        if form.is_valid():
            form.save()
            _log_activity(request.user, ActivityLog.UPDATE, "Transaction", _transaction_repr(tx, request.user.currency))
            messages.success(request, "Transaction updated.")
            return redirect("transactions")
    else:
        form = TransactionForm(instance=tx, user=request.user)

    return render(request, "transaction_form.html", {"form": form, "tx": tx})


@login_required
def transaction_delete_view(request, pk):
    tx = get_object_or_404(Transaction, pk=pk, user=request.user)

    if request.method == "POST":
        object_repr = _transaction_repr(tx, request.user.currency)
        tx.delete()
        _log_activity(request.user, ActivityLog.DELETE, "Transaction", object_repr)
        messages.success(request, "Transaction deleted.")
        return redirect("transactions")

    return render(request, "transaction_confirm_delete.html", {"tx": tx})


def _month_start(d: date) -> date:
    return d.replace(day=1)


def _month_end(d: date) -> date:
    last_day = monthrange(d.year, d.month)[1]
    return d.replace(day=last_day)


def _budget_status_rows(user, month_start, month_end):
    """Spend vs. limit per category for one month, with over/near-limit
    flags. Shared by the budgets page (all categories) and the dashboard
    proactive alert (only categories that need attention)."""
    categories = Category.objects.filter(user=user).order_by("name")

    spent_map = {
        row["category_id"]: (row["total"] or Decimal("0"))
        for row in (
            Transaction.objects.filter(
                user=user,
                type=Transaction.EXPENSE,
                date__gte=month_start,
                date__lte=month_end,
            )
            .values("category_id")
            .annotate(total=Sum("amount"))
        )
    }

    limits = BudgetLimit.objects.filter(user=user, month=month_start).select_related("category")
    limit_map = {b.category_id: b for b in limits}

    rows = []
    for c in categories:
        spent = spent_map.get(c.id, Decimal("0"))
        b = limit_map.get(c.id)
        limit_val = b.limit if b else None

        if limit_val and limit_val > 0:
            pct = (spent / limit_val) * 100
            pct_int = int(min(pct, 999))  # cap for UI
        else:
            pct_int = 0

        is_over = limit_val is not None and spent > limit_val
        rows.append({
            "category": c,
            "spent": spent,
            "limit": limit_val,
            "pct": pct_int,
            "budget": b,
            "is_over": is_over,
            "is_near": (not is_over) and limit_val is not None and limit_val > 0 and pct_int >= 80,
        })

    return rows


@login_required
def budgets_view(request):
    month_str = request.GET.get("month", "").strip()
    if month_str:
        try:
            selected = date.fromisoformat(month_str)
        except ValueError:
            selected = date.today()
    else:
        selected = date.today()

    month_start = _month_start(selected)
    month_end = _month_end(selected)

    rows = _budget_status_rows(request.user, month_start, month_end)

    total_spent = sum((r["spent"] for r in rows), Decimal("0"))
    total_limit = sum((r["limit"] for r in rows if r["limit"] is not None), Decimal("0"))

    return render(request, "budgets.html", {
        "rows": rows,
        "month": month_start,
        "currency": request.user.currency,
        "total_spent": total_spent,
        "total_limit": total_limit,
    })


@login_required
def budget_add_view(request):
    if request.method == "POST":
        form = BudgetLimitForm(request.POST, user=request.user)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.user = request.user
            obj.month = _month_start(obj.month)  # normalize
            obj.save()
            _log_activity(request.user, ActivityLog.CREATE, "Budget Limit", _budget_limit_repr(obj, request.user.currency))
            messages.success(request, "Budget limit saved.")
            return redirect("budgets")
    else:
        form = BudgetLimitForm(user=request.user, initial={"month": date.today().replace(day=1)})

    return render(request, "budget_form.html", {"form": form, "mode": "add"})


@login_required
def budget_edit_view(request, pk):
    b = get_object_or_404(BudgetLimit, pk=pk, user=request.user)

    if request.method == "POST":
        form = BudgetLimitForm(request.POST, instance=b, user=request.user)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.month = _month_start(obj.month)
            obj.save()
            _log_activity(request.user, ActivityLog.UPDATE, "Budget Limit", _budget_limit_repr(obj, request.user.currency))
            messages.success(request, "Budget limit updated.")
            return redirect("budgets")
    else:
        form = BudgetLimitForm(instance=b, user=request.user)

    return render(request, "budget_form.html", {"form": form, "mode": "edit", "budget": b})


@login_required
def budget_delete_view(request, pk):
    b = get_object_or_404(BudgetLimit, pk=pk, user=request.user)

    if request.method == "POST":
        object_repr = _budget_limit_repr(b, request.user.currency)
        b.delete()
        _log_activity(request.user, ActivityLog.DELETE, "Budget Limit", object_repr)
        messages.success(request, "Budget limit deleted.")
        return redirect("budgets")

    return render(request, "budget_confirm_delete.html", {"budget": b})


@login_required
def charts_view(request):
    today = date.today()
    month_start = today.replace(day=1)
    last_day = monthrange(today.year, today.month)[1]
    month_end = today.replace(day=last_day)

    pie_rows = (
        Transaction.objects.filter(
            user=request.user,
            type=Transaction.EXPENSE,
            date__gte=month_start,
            date__lte=month_end,
        )
        .values("category__name", "category__color", "category__icon")
        .annotate(total=Sum("amount"))
        .order_by("-total")
    )

    pie_labels, pie_values, pie_colors = [], [], []
    for r in pie_rows:
        icon = r["category__icon"] or "🏷️"
        name = r["category__name"] or "Other"
        pie_labels.append(f"{icon} {name}".strip())
        pie_values.append(float(r["total"] or 0))
        pie_colors.append(r["category__color"] or "#6366f1")

    monthly = (
        Transaction.objects.filter(user=request.user)
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(
            income=Sum("amount", filter=models.Q(type=Transaction.INCOME)),
            expense=Sum("amount", filter=models.Q(type=Transaction.EXPENSE)),
        )
        .order_by("month")
    )

    chart_labels, chart_income, chart_expense = [], [], []
    for m in monthly:
        if m["month"]:
            chart_labels.append(m["month"].strftime("%b %Y"))
            chart_income.append(float(m["income"] or 0))
            chart_expense.append(float(m["expense"] or 0))

    return render(request, "charts.html", {
        "currency": request.user.currency,
        "month_label": month_start.strftime("%b %Y"),
        "pie_labels": pie_labels,
        "pie_values": pie_values,
        "pie_colors": pie_colors,
        "chart_labels": chart_labels,
        "chart_income": chart_income,
        "chart_expense": chart_expense,
    })


@login_required
def compare_view(request):
    available_years = [
        d.year for d in Transaction.objects.filter(user=request.user).dates("date", "year", order="DESC")
    ]

    year_str = request.GET.get("year", "").strip()
    try:
        selected_year = int(year_str) if year_str else (available_years[0] if available_years else date.today().year)
    except ValueError:
        selected_year = available_years[0] if available_years else date.today().year

    if selected_year not in available_years:
        available_years = sorted(set(available_years) | {selected_year}, reverse=True)

    monthly = (
        Transaction.objects.filter(user=request.user, date__year=selected_year)
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(
            income=Sum("amount", filter=models.Q(type=Transaction.INCOME)),
            expense=Sum("amount", filter=models.Q(type=Transaction.EXPENSE)),
        )
    )
    month_map = {m["month"].month: m for m in monthly if m["month"]}

    raw_rows = []
    for month_num in range(1, 13):
        data = month_map.get(month_num)
        income = (data["income"] or Decimal("0")) if data else Decimal("0")
        expense = (data["expense"] or Decimal("0")) if data else Decimal("0")
        raw_rows.append({
            "label": date(selected_year, month_num, 1).strftime("%b"),
            "income": income,
            "expense": expense,
            "balance": income - expense,
        })

    year_total_income = sum((r["income"] for r in raw_rows), Decimal("0"))
    year_total_expense = sum((r["expense"] for r in raw_rows), Decimal("0"))

    max_amount = max([r["income"] for r in raw_rows] + [r["expense"] for r in raw_rows] + [Decimal("1")])
    month_rows = []
    for r in raw_rows:
        month_rows.append({
            **r,
            "income_pct": int(min(r["income"] / max_amount * 100, 100)),
            "expense_pct": int(min(r["expense"] / max_amount * 100, 100)),
        })

    yearly = (
        Transaction.objects.filter(user=request.user)
        .annotate(year=TruncYear("date"))
        .values("year")
        .annotate(
            income=Sum("amount", filter=models.Q(type=Transaction.INCOME)),
            expense=Sum("amount", filter=models.Q(type=Transaction.EXPENSE)),
        )
        .order_by("-year")
    )
    year_rows = []
    for y in yearly:
        income = y["income"] or Decimal("0")
        expense = y["expense"] or Decimal("0")
        year_rows.append({
            "year": y["year"].year,
            "income": income,
            "expense": expense,
            "balance": income - expense,
        })

    return render(request, "compare.html", {
        "currency": request.user.currency,
        "available_years": available_years,
        "selected_year": selected_year,
        "month_rows": month_rows,
        "year_total_income": year_total_income,
        "year_total_expense": year_total_expense,
        "year_total_balance": year_total_income - year_total_expense,
        "year_rows": year_rows,
    })


def _savings_goal_repr(goal, currency):
    return f"{goal.name} · target {goal.target_amount} {currency}"


@login_required
def savings_goals_view(request):
    goals = SavingsGoal.objects.filter(user=request.user)

    celebrate_raw = request.GET.get("celebrate", "").strip()
    celebrate_id = int(celebrate_raw) if celebrate_raw.isdigit() else None

    rows = []
    for g in goals:
        pct = int(min((g.saved_amount / g.target_amount) * 100, 100)) if g.target_amount else 0
        rows.append({
            "goal": g,
            "pct": pct,
            "remaining": max(g.target_amount - g.saved_amount, Decimal("0")),
            "is_complete": g.saved_amount >= g.target_amount,
        })

    return render(request, "savings_goals.html", {
        "rows": rows,
        "currency": request.user.currency,
        "celebrate_id": celebrate_id,
    })


@login_required
def savings_goal_add_view(request):
    if request.method == "POST":
        form = SavingsGoalForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.user = request.user
            obj.save()
            _log_activity(request.user, ActivityLog.CREATE, "Savings Goal", _savings_goal_repr(obj, request.user.currency))
            messages.success(request, "Savings goal created!")
            return redirect("savings_goals")
    else:
        form = SavingsGoalForm()

    return render(request, "savings_goal_form.html", {"form": form, "mode": "add"})


@login_required
def savings_goal_edit_view(request, pk):
    goal = get_object_or_404(SavingsGoal, pk=pk, user=request.user)

    if request.method == "POST":
        form = SavingsGoalForm(request.POST, instance=goal)
        if form.is_valid():
            form.save()
            _log_activity(request.user, ActivityLog.UPDATE, "Savings Goal", _savings_goal_repr(goal, request.user.currency))
            messages.success(request, "Savings goal updated.")
            return redirect("savings_goals")
    else:
        form = SavingsGoalForm(instance=goal)

    return render(request, "savings_goal_form.html", {"form": form, "mode": "edit", "goal": goal})


@login_required
def savings_goal_add_funds_view(request, pk):
    goal = get_object_or_404(SavingsGoal, pk=pk, user=request.user)

    if request.method == "POST":
        amount = _parse_amount(request.POST.get("amount"))
        if amount is None:
            messages.error(request, "Please enter a valid amount greater than zero.")
        else:
            was_complete = goal.saved_amount >= goal.target_amount
            goal.saved_amount = goal.saved_amount + amount
            goal.save()
            _log_activity(
                request.user, ActivityLog.UPDATE, "Savings Goal",
                f"{goal.name} · added {amount} {request.user.currency}",
            )

            just_completed = (not was_complete) and goal.saved_amount >= goal.target_amount
            if just_completed:
                messages.success(request, f"🎉 Goal reached! You hit your target for \"{goal.name}\".")
                return redirect(f"{reverse('savings_goals')}?celebrate={goal.pk}")

            messages.success(request, "Funds added to your goal.")

    return redirect("savings_goals")


@login_required
def savings_goal_delete_view(request, pk):
    goal = get_object_or_404(SavingsGoal, pk=pk, user=request.user)

    if request.method == "POST":
        object_repr = _savings_goal_repr(goal, request.user.currency)
        goal.delete()
        _log_activity(request.user, ActivityLog.DELETE, "Savings Goal", object_repr)
        messages.success(request, "Savings goal deleted.")
        return redirect("savings_goals")

    return render(request, "savings_goal_confirm_delete.html", {"goal": goal})


@login_required
def activity_log_view(request):
    logs = ActivityLog.objects.filter(user=request.user)
    paginator = Paginator(logs, 50)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "activity_log.html", {"page_obj": page_obj})


def _recurring_repr(r, currency):
    return f"{r.get_type_display()} · {r.category.name} · {r.amount} {currency} · day {r.day_of_month}"


@login_required
def recurring_list_view(request):
    items = RecurringTransaction.objects.filter(user=request.user).select_related("category", "wallet")
    return render(request, "recurring_transactions.html", {"items": items, "currency": request.user.currency})


@login_required
def recurring_add_view(request):
    categories = Category.objects.filter(user=request.user).order_by("name")
    wallets = Wallet.objects.filter(user=request.user).order_by("name")

    if not categories.exists() or not wallets.exists():
        return render(request, "recurring_transaction_form.html", {
            "mode": "add",
            "missing_categories": not categories.exists(),
            "missing_wallets": not wallets.exists(),
        })

    if request.method == "POST":
        form = RecurringTransactionForm(request.POST, user=request.user)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.user = request.user
            obj.save()
            _log_activity(request.user, ActivityLog.CREATE, "Recurring Transaction", _recurring_repr(obj, request.user.currency))
            messages.success(request, "Recurring transaction created!")
            return redirect("recurring_transactions")
    else:
        form = RecurringTransactionForm(user=request.user, initial={"start_date": date.today(), "day_of_month": 1})

    return render(request, "recurring_transaction_form.html", {"form": form, "mode": "add"})


@login_required
def recurring_edit_view(request, pk):
    r = get_object_or_404(RecurringTransaction, pk=pk, user=request.user)

    if request.method == "POST":
        form = RecurringTransactionForm(request.POST, instance=r, user=request.user)
        if form.is_valid():
            form.save()
            _log_activity(request.user, ActivityLog.UPDATE, "Recurring Transaction", _recurring_repr(r, request.user.currency))
            messages.success(request, "Recurring transaction updated.")
            return redirect("recurring_transactions")
    else:
        form = RecurringTransactionForm(instance=r, user=request.user)

    return render(request, "recurring_transaction_form.html", {"form": form, "mode": "edit", "item": r})


@login_required
def recurring_delete_view(request, pk):
    r = get_object_or_404(RecurringTransaction, pk=pk, user=request.user)

    if request.method == "POST":
        object_repr = _recurring_repr(r, request.user.currency)
        r.delete()
        _log_activity(request.user, ActivityLog.DELETE, "Recurring Transaction", object_repr)
        messages.success(request, "Recurring transaction deleted.")
        return redirect("recurring_transactions")

    return render(request, "recurring_transaction_confirm_delete.html", {"item": r})
